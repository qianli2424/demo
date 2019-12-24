#!/usr/bin/python3
# _*_ coding:utf-8 _*_
# @Author:啊志
# 2019/12/23/15:47
# @Email:1427245437@qq.com
import openpyxl
feil=openpyxl.load_workbook('login.xlsx')
sheet1=feil['login']
max_row=sheet1.max_row
max_col=sheet1.max_column

#获取用户信息列表
user_list={}
#读取表格第二列用户名
for x in range(3,max_row+1):
    key=sheet1.cell(row=x,column=1).value
    # print(key)
#读取表格中用户信息
    info_list=[]
    for y in range(2,max_col):
        info=sheet1.cell(row=x,column=y).value
        info_list.append(str(info))
        user_list[key]=info_list
print(user_list)
print("----------------------------登录系统------------------------------")
while True:
    name=input("请输入用户名")
    if name==" ":
        print("用户名不能为空")
    elif name not in user_list.keys():
        print("用户名不存在")

    else:
        fail_count=int(user_list[name][2])
        while fail_count<3:
            passwd=input("请输入密码")
            if user_list[name][0] == passwd:
                print("登录成功")
                break
            elif passwd==" ":
                print("密码不能为空")

            elif user_list[name][2]==3:
                print("密码超出错误次数，被锁定")
                user_list[name][1]=""
                print(user_list)
                break

            elif user_list[name][0] != passwd:
                print("密码错误")
                fail_count += 1
                print("你还有{}次机会".format(3 -fail_count))
                # print(users)
            # elif users[name][2]==0:
            #     print("")
        break

